import os
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

# Load environment variables from .env file
load_dotenv()

# newnew

# Shopify GraphQL API credentials
shop_name = os.getenv('SHOP_NAME')
admin_api_token = os.getenv('API_ACCESS_TOKEN')
export_path = os.getenv('EXPORT_PAYMENTS_PATH')

# Global date variables
start_date = "2024-10-01"  # Replace with your desired start date
end_date = "2024-10-31"  # Replace with your desired end date


def fetch_payouts():
    all_payouts = []
    has_next_page = True
    after_cursor = None

    while has_next_page:
        after_clause = f', after: "{after_cursor}"' if after_cursor else ""
        query = f'''
        {{
          shopifyPaymentsAccount {{
            payouts(first: 100{after_clause}) {{
              edges {{
                node {{
                  id
                  issuedAt
                  net {{
                    amount
                    currencyCode
                  }}
                  gross {{
                    amount
                  }}
                  status
                  summary {{
                    adjustmentsGross {{
                      amount
                    }}
                    chargesGross {{
                      amount
                    }}
                    refundsFeeGross {{
                      amount
                    }}
                    reservedFundsGross {{
                      amount
                    }}
                    chargesFee {{
                      amount
                    }}
                    retriedPayoutsGross {{
                      amount
                    }}
                  }}
                }}
                cursor
              }}
              pageInfo {{
                hasNextPage
                endCursor
              }}
            }}
          }}
        }}
        '''

        url = f"https://{shop_name}.myshopify.com/admin/api/2024-10/graphql.json"
        headers = {
            "Content-Type": "application/json",
            "X-Shopify-Access-Token": admin_api_token
        }

        response = requests.post(url, json={"query": query}, headers=headers)
        response_data = response.json()

        if 'data' not in response_data:
            raise Exception("No data returned from Shopify.")

        payouts = response_data['data']['shopifyPaymentsAccount']['payouts']['edges']
        all_payouts.extend(payouts)

        page_info = response_data['data']['shopifyPaymentsAccount']['payouts']['pageInfo']
        has_next_page = page_info['hasNextPage']
        after_cursor = page_info['endCursor']

    filtered_payouts = []
    for payout in all_payouts:
        payout_date = datetime.strptime(
            payout['node']['issuedAt'], "%Y-%m-%dT%H:%M:%S%z").date()
        if datetime.strptime(start_date, "%Y-%m-%d").date() <= payout_date <= datetime.strptime(end_date, "%Y-%m-%d").date():
            filtered_payouts.append(payout)

    if not filtered_payouts:
        print(f"No payouts found between {start_date} and {end_date}.")

    return filtered_payouts


def fetch_balance_transactions(payout_id):
    transactions = []
    has_next_page = True
    after_cursor = None

    while has_next_page:
        after_clause = f', after: "{after_cursor}"' if after_cursor else ""
        query = f'''
        {{
          shopifyPaymentsAccount {{
            balanceTransactions(first: 100{after_clause}) {{
              edges {{
                node {{
                  id
                  transactionDate
                  type
                  associatedOrder {{
                    name
                  }}
                  amount {{
                    amount
                  }}
                  fee {{
                    amount
                  }}
                  net {{
                    amount
                  }}
                  associatedPayout {{
                    id
                  }}
                }}
                cursor
              }}
              pageInfo {{
                hasNextPage
                endCursor
              }}
            }}
          }}
        }}
        '''
        url = f"https://{shop_name}.myshopify.com/admin/api/2024-10/graphql.json"
        headers = {
            "Content-Type": "application/json",
            "X-Shopify-Access-Token": admin_api_token
        }

        response = requests.post(url, json={"query": query}, headers=headers)
        response_data = response.json()

        if 'data' not in response_data or not response_data['data']['shopifyPaymentsAccount']:
            raise Exception(
                f"No transaction data returned for payout {payout_id}. Response: {response_data}")

        balance_transactions = response_data['data']['shopifyPaymentsAccount']['balanceTransactions'].get(
            'edges', [])

        for transaction in balance_transactions:
            if transaction['node']['associatedPayout'] and transaction['node']['associatedPayout']['id'] == payout_id:
                transactions.append(transaction['node'])

        page_info = response_data['data']['shopifyPaymentsAccount']['balanceTransactions'].get(
            'pageInfo', {})
        has_next_page = page_info.get('hasNextPage', False)
        after_cursor = page_info.get('endCursor', None)

    return transactions


def fetch_order_id(order_number):
    """
    Fetches the global ID of the order based on the order number.
    """
    query = f'''
    {{
      orders(first: 1, query: "name:{order_number}") {{
        edges {{
          node {{
            id
            name
          }}
        }}
      }}
    }}
    '''

    url = f"https://{shop_name}.myshopify.com/admin/api/2024-10/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": admin_api_token
    }

    response = requests.post(url, json={"query": query}, headers=headers)
    if response.status_code == 200:
        response_data = response.json()
        if response_data['data']['orders']['edges']:
            global_order_id = response_data['data']['orders']['edges'][0]['node']['id']
            # print(f"Fetched Global Order ID: {global_order_id}")
            return global_order_id
        else:
            print(f"No order found for {order_number}")
            return None
    else:
        print(
            f"Error fetching order ID: {response.status_code}, {response.text}")
        return None


def fetch_order_refund_details(order_id):
    """
    Fetches detailed refund and transaction information for a specific order using the correct global order ID.
    """
    # print(f"Order ID passed to refund query: {order_id}")

    query = f'''
    {{
      order(id: "{order_id}") {{
        returns(first: 10) {{
          edges {{
            node {{
              id
              returnShippingFees {{
                amountSet {{
                  shopMoney {{
                    amount
                  }}
                }}
              }}
              refunds(first: 10) {{
                edges {{
                  node {{
                    id
                    refundLineItems(first: 10) {{
                      edges {{
                        node {{
                          quantity
                          lineItem {{
                            variant {{
                              barcode
                            }}
                            sku
                          }}
                          subtotal
                          totalTax
                        }}
                      }}
                    }}
                    refundShippingLines(first: 10) {{
                      edges {{
                        node {{
                          subtotalAmountSet {{
                            shopMoney {{
                              amount
                            }}
                          }}
                          taxAmountSet {{
                            shopMoney {{
                              amount
                            }}
                          }}
                        }}
                      }}
                    }}
                    totalRefunded {{
                      amount
                    }}
                  }}
                }}
              }}
            }}
          }}
        }}
        refunds {{
          id
          createdAt
          totalRefundedSet {{
            shopMoney {{
              amount
              currencyCode
            }}
          }}
          refundLineItems(first: 10) {{
            edges {{
              node {{
                quantity
                lineItem {{
                  variant {{
                    barcode
                  }}
                  sku
                }}
                subtotalSet {{
                  shopMoney {{
                    amount
                    currencyCode
                  }}
                }}
                totalTaxSet {{
                  shopMoney {{
                    amount
                    currencyCode
                  }}
                }}
              }}
            }}
          }}
          refundShippingLines(first: 10) {{
            edges {{
              node {{
                subtotalAmountSet {{
                  shopMoney {{
                    amount
                  }}
                }}
                taxAmountSet {{
                  shopMoney {{
                    amount
                  }}
                }}
              }}
            }}
          }}
        }}
        transactions(first: 10) {{
          amount
          createdAt
          kind
          status
          totalUnsettled
          errorCode
          maximumRefundable
          multiCapturable
        }}
      }}
    }}
    '''

    url = f"https://{shop_name}.myshopify.com/admin/api/2024-10/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": admin_api_token
    }

    response = requests.post(url, json={"query": query}, headers=headers)
    if response.status_code == 200:
        response_data = response.json()
        # print(f"Response received for order {order_id}: {response_data}")

        if response_data.get('data', {}).get('order', {}):
            return response_data
        else:
            print(f"No refund or transaction data found for order {order_id}")
            return None
    else:
        print(f"Error: {response.status_code}, {response.text}")
        return None


def process_refunds(order_number):
    """
    Fetches the global order ID and refund details for a given order number.
    """
    global_order_id = fetch_order_id(order_number)
    if global_order_id:
        refund_data = fetch_order_refund_details(global_order_id)
        # if refund_data:
        #     print(f"Refund Data for order {order_number}: {refund_data}")
        # else:
        #     print(f"No refund data found for order {order_number}")
    else:
        print(f"Unable to fetch refund data for order {order_number}")


def create_excel(payouts, output_file):
    workbook = Workbook()

    # Sheet 1: Payout Overview
    sheet1 = workbook.active
    sheet1.title = "Payout Overview"

    headers = ["Payout Date", "Payout ID", "Status", "Charges", "Refunds", "Adjustments",
               "Reserved Funds", "Fees", "Retired Amount", "Total", "Currency"]
    sheet1.append(headers)

    # Create the Refund Details sheet with updated headers
    refund_sheet = workbook.create_sheet(title="Refund Details")
    refund_sheet.append(["Order Number", "Payout ID", "Refund Date", "Refund Amount(T)", "UPC", "SKU", "QTY", "Subtotal",
                         "Tax", "Shipping(T)", "Shipping Tax(T)", "Return Shipping Fee(T)", "Transaction ID"])

    refund_rows = []

    for payout in payouts:
        node = payout['node']
        summary = node['summary']
        payout_id = node['id'][-10:]

        # Create the hyperlink for the Payout Date, linking to the Payout Details sheet
        payout_date = datetime.strptime(
            node["issuedAt"], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y")
        hyperlink = f'=HYPERLINK("#\'Payout {payout_id}\'!A1", "{payout_date}")'

        # Append the row with the hyperlink for Payout Date in the Payout Overview sheet
        row = [
            hyperlink,  # Hyperlink in the Payout Date column
            payout_id,
            node['status'],
            float(summary['chargesGross']['amount']
                  ) if summary['chargesGross'] else 0.00,
            float(summary['refundsFeeGross']['amount']
                  ) if summary['refundsFeeGross'] else 0.00,
            float(summary['adjustmentsGross']['amount']
                  ) if summary['adjustmentsGross'] else 0.00,
            float(summary['reservedFundsGross']['amount']
                  ) if summary['reservedFundsGross'] else 0.00,
            float(summary['chargesFee']['amount']
                  ) if summary['chargesFee'] else 0.00,
            float(summary['retriedPayoutsGross']['amount']
                  ) if summary['retriedPayoutsGross'] else 0.00,
            float(node['net']['amount']),
            node['net']['currencyCode']
        ]
        sheet1.append(row)

        # Style the hyperlink in blue and underlined
        cell = sheet1.cell(row=sheet1.max_row, column=1)
        cell.font = Font(color="0000FF", underline="single")

        transactions = fetch_balance_transactions(node['id'])

        # Create Payout Details sheet for this payout
        clean_title = f"Payout {payout_id}"
        payout_sheet = workbook.create_sheet(title=clean_title)

        # Add "Back to Overview" link in C1
        payout_sheet.cell(
            row=1, column=3, value='=HYPERLINK("#\'Payout Overview\'!A1", "Back to Overview")')
        cell = payout_sheet.cell(row=1, column=3)
        cell.font = Font(color="0000FF", underline="single")

        # Add payout summary info at the top of the Payout Details sheet
        payout_sheet.append(["Payout Date:", datetime.strptime(
            node['issuedAt'], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y")])
        payout_sheet.append(["Payout ID:", payout_id])
        payout_sheet.append(
            ["Total:", f"{node['net']['amount']} {node['net']['currencyCode']}"])
        payout_sheet.append([])  # Blank row

        # Charges, Refunds, Adjustments with total calculation in the Payout Details sheet
        payout_sheet.append(["", "Gross", "Fees", "Total"])
        row_idx = payout_sheet.max_row + 1  # To track the row index

        # Add the rows for Charges, Refunds, and Adjustments
        payout_sheet.append(["Charges", summary['chargesGross']['amount'],
                            summary['chargesFee']['amount'],
                            float(summary['chargesGross']['amount']) - float(summary['chargesFee']['amount'])])
        payout_sheet.append(["Refunds", summary['refundsFeeGross']['amount'],
                            "0.00", summary['refundsFeeGross']['amount']])
        payout_sheet.append(["Adjustments", summary['adjustmentsGross']['amount'],
                            "0.00", summary['adjustmentsGross']['amount']])
        payout_sheet.append([])  # Blank row

        # Format the cells for Charges, Refunds, and Adjustments
        for row in range(row_idx, row_idx + 3):  # 3 rows: Charges, Refunds, Adjustments
            for col in range(2, 5):  # Columns B, C, D (2, 3, 4 in 1-based indexing)
                cell = payout_sheet.cell(row=row, column=col)
                cell.number_format = '#,##0.00'
                # Convert string values to float before formatting
                if isinstance(cell.value, str):
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass  # Keep original value if conversion fails

        # Add transaction details to the Payout Details sheet
        payout_sheet.append(["Transaction Date", "Type", "Order", "Amount",
                            "Fee", "Net", "Currency", "Transaction ID"])  # Added Transaction ID

        for transaction in transactions:
            processed_refund_ids = set()  # Reset for each transaction line
            specific_match_found = False  # Initialize to False for each transaction

            # Skip TRANSFER type transactions
            if transaction['type'] == "TRANSFER":
                continue

            transaction_row = [
                datetime.strptime(
                    transaction['transactionDate'], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y"),
                transaction['type'],
                transaction['associatedOrder']['name'].replace(
                    "#", "") if transaction['associatedOrder'] else "N/A",
                float(transaction['amount']['amount']
                      ) if transaction['amount'] else 0.00,
                float(transaction['fee']['amount']
                      ) if transaction['fee'] else 0.00,
                float(transaction['net']['amount']
                      ) if transaction['net'] else 0.00,
                node['net']['currencyCode'],
                transaction['id'].replace(
                    'gid://shopify/ShopifyPaymentsBalanceTransaction/', '')
            ]
            payout_sheet.append(transaction_row)

            # If this is a REFUND transaction, fetch the refund details and store them in the refund sheet

            if transaction['type'] != "REFUND":
                continue

            # Extract the order number and transaction amount
            order_number = transaction['associatedOrder']['name'].replace(
                "#", "")
            transaction_amount = float(
                transaction['amount']['amount']) if transaction['amount'] else 0.00
            print(
                f"Processing REFUND transaction for order {order_number} with amount {transaction_amount}...")

            # Step 1: Fetch the global order ID
            global_order_id = fetch_order_id(order_number)

            if global_order_id:
                # Fetch refund and transaction details for this order
                refund_response = fetch_order_refund_details(global_order_id)
                if refund_response and 'data' in refund_response:
                    order_data = refund_response['data']['order']
                    refund_processed = False  # Add this flag to track if we've processed a refund

                    # Check if refund exists in returns section
                    if order_data.get('returns'):
                        for return_data in order_data['returns']['edges']:
                            return_node = return_data['node']
                            return_refunds = return_node['refunds']['edges']
                            for return_refund in return_refunds:
                                refund_id = return_refund['node']['id']
                                refund_amount = float(
                                    return_refund['node']['totalRefunded']['amount'])

                                if abs(refund_amount - abs(transaction_amount)) < 0.01:
                                    # Get shipping lines from the refund
                                    shipping_lines = return_refund['node']['refundShippingLines']['edges']

                                    # Calculate shipping and tax amounts
                                    shipping_refund = sum(
                                        float(
                                            edge['node']['subtotalAmountSet']['shopMoney']['amount'])
                                        for edge in shipping_lines
                                    ) if shipping_lines else 0.00

                                    shipping_tax = sum(
                                        float(edge['node']['taxAmountSet']
                                              ['shopMoney']['amount'])
                                        for edge in shipping_lines
                                    ) if shipping_lines else 0.00

                                    # Get return shipping fee if it exists
                                    return_shipping_fee = float(
                                        return_node['returnShippingFees'][0]['amountSet']['shopMoney']['amount']
                                    ) if return_node['returnShippingFees'] else 0.00

                                    # Process line items and shipping lines
                                    for index, line_item in enumerate(return_refund['node']['refundLineItems']['edges']):
                                        line_node = line_item['node']
                                        refund_row = [
                                            order_number,
                                            f'=HYPERLINK("#\'Payout {payout_id}\'!A1", "{payout_id}")',
                                            datetime.strptime(
                                                transaction['transactionDate'], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y"),
                                            -abs(refund_amount) if index == 0 else 0.00,
                                            line_node['lineItem']['variant'].get(
                                                'barcode', 'N/A') if line_node['lineItem'].get('variant') else 'N/A',
                                            line_node['lineItem'].get(
                                                'sku', 'N/A'),
                                            line_node['quantity'],
                                            -abs(float(line_node['subtotal'])),
                                            -abs(float(line_node['totalTax'])),
                                            # Changed to use shipping_refund
                                            -abs(shipping_refund) if index == 0 else 0.00,
                                            # Changed to use shipping_tax
                                            -abs(shipping_tax) if index == 0 else 0.00,
                                            return_shipping_fee if index == 0 else 0.00,
                                            transaction['id'].replace(
                                                'gid://shopify/ShopifyPaymentsBalanceTransaction/', '')
                                        ]
                                        refund_sheet.append(refund_row)
                                        cell = refund_sheet.cell(
                                            row=refund_sheet.max_row, column=2)
                                        cell.font = Font(
                                            color="0000FF", underline="single")
                                    refund_processed = True  # Set the flag to indicate we've processed this refund
                                    break  # Break out of return_refunds loop
                            if refund_processed:
                                break  # Break out of returns loop if we've processed a refund

                    # Only proceed with checking regular refunds if we haven't processed a return refund
                    if not refund_processed:
                        # Step 2: Try matching the transaction amount with order refunds
                        if order_data.get('refunds'):
                            for refund in order_data['refunds']:
                                refund_id = refund['id']
                                refund_amount = float(
                                    refund['totalRefundedSet']['shopMoney']['amount'])

                                if abs(refund_amount - abs(transaction_amount)) < 0.01:
                                    # Process this refund's line items and shipping lines
                                    refund_line_items = refund.get(
                                        'refundLineItems', {}).get('edges', [])
                                    shipping_lines = refund.get(
                                        'refundShippingLines', {}).get('edges', [])
                                    total_refund = -abs(refund_amount)

                                    # Handle manual refund case
                                    if not refund_line_items and not shipping_lines:
                                        refund_row = [
                                            order_number,
                                            f'=HYPERLINK("#\'Payout {payout_id}\'!A1", "{payout_id}")',
                                            datetime.strptime(
                                                refund['createdAt'], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y"),
                                            total_refund,
                                            "Manual Refund",
                                            "N/A",
                                            0,
                                            total_refund,
                                            0.00,
                                            0.00,
                                            0.00,
                                            0.00,
                                            transaction['id'].replace(
                                                'gid://shopify/ShopifyPaymentsBalanceTransaction/', '')
                                        ]
                                        refund_sheet.append(refund_row)
                                        cell = refund_sheet.cell(
                                            row=refund_sheet.max_row, column=2)
                                        cell.font = Font(
                                            color="0000FF", underline="single")
                                        refund_processed = True
                                        continue

                                    # Calculate shipping and tax amounts
                                    shipping_refund = sum(
                                        float(
                                            edge['node']['subtotalAmountSet']['shopMoney']['amount'])
                                        for edge in shipping_lines
                                    ) if shipping_lines else 0.00

                                    shipping_tax = sum(
                                        float(edge['node']['taxAmountSet']
                                              ['shopMoney']['amount'])
                                        for edge in shipping_lines
                                    ) if shipping_lines else 0.00

                                    # Process line items
                                    for index, line_item in enumerate(refund_line_items):
                                        line_node = line_item['node']
                                        refund_row = [
                                            order_number,
                                            f'=HYPERLINK("#\'Payout {payout_id}\'!A1", "{payout_id}")',
                                            datetime.strptime(
                                                refund['createdAt'], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y"),
                                            total_refund if index == 0 else 0.00,
                                            line_node['lineItem']['variant'].get(
                                                'barcode', 'N/A') if line_node['lineItem'].get('variant') else 'N/A',
                                            line_node['lineItem'].get(
                                                'sku', 'N/A'),
                                            line_node['quantity'],
                                            -abs(float(line_node['subtotalSet']['shopMoney']['amount'])),
                                            -abs(float(line_node['totalTaxSet']['shopMoney']['amount'])),
                                            -abs(shipping_refund) if index == 0 else 0.00,
                                            -abs(shipping_tax) if index == 0 else 0.00,
                                            0.00,
                                            transaction['id'].replace(
                                                'gid://shopify/ShopifyPaymentsBalanceTransaction/', '')
                                        ]
                                        refund_sheet.append(refund_row)
                                        cell = refund_sheet.cell(
                                            row=refund_sheet.max_row, column=2)
                                        cell.font = Font(
                                            color="0000FF", underline="single")
                                    refund_processed = True
                                    break

                        # Step 3: Only check for AUTH/CAPTURE difference if we haven't processed any refund yet
                        if not refund_processed and order_data.get('transactions'):
                            auth_amount = None
                            capture_amount = None

                            for txn in order_data['transactions']:
                                if txn['kind'] == "AUTHORIZATION" and txn['status'] == "SUCCESS":
                                    auth_amount = float(txn['amount'])
                                elif txn['kind'] == "CAPTURE" and txn['status'] == "SUCCESS":
                                    capture_amount = float(txn['amount'])

                            # Calculate difference and match it with the transaction amount
                            if auth_amount is not None and capture_amount is not None:
                                difference = auth_amount - capture_amount
                                if abs(difference - abs(transaction_amount)) < 0.01:
                                    # Match found in auth/capture difference
                                    refund_row = [
                                        order_number,
                                        f'=HYPERLINK("#\'Payout {payout_id}\'!A1", "{payout_id}")',
                                        datetime.strptime(
                                            transaction['transactionDate'], "%Y-%m-%dT%H:%M:%S%z").strftime("%m/%d/%Y"),
                                        -abs(transaction_amount),
                                        "Auth Release",
                                        "N/A",
                                        0,
                                        -abs(transaction_amount),
                                        0.00,
                                        0.00,
                                        0.00,
                                        0.00,
                                        transaction['id'].replace(
                                            'gid://shopify/ShopifyPaymentsBalanceTransaction/', '')
                                    ]
                                    refund_sheet.append(refund_row)
                                    cell = refund_sheet.cell(
                                        row=refund_sheet.max_row, column=2)
                                    cell.font = Font(
                                        color="0000FF", underline="single")

    sorted_refund_rows = sorted(
        refund_rows, key=lambda x: x[0])  # Sort by Order Number
    for row in sorted_refund_rows:
        refund_sheet.append(row)
        # Style the hyperlink
        cell = refund_sheet.cell(row=refund_sheet.max_row, column=2)
        cell.font = Font(color="0000FF", underline="single")

    # Format all number fields as numbers in the Refund Details sheet
    for row in refund_sheet.iter_rows(min_row=2):
        for column in range(3, 13):  # Updated range to include new column
            cell = row[column - 1]
            if column == 7:  # Column G (QTY)
                cell.number_format = '0'
            else:
                cell.number_format = '#,##0.00'

    # Create Payout Comparison sheet
    comparison_sheet = workbook.create_sheet(title="Payout Comparison")
    comparison_sheet.append(
        ["Order", "Original Amount", "Auth Release Amount", "Total Amount"])

    # Dictionary to store all charges
    charges_dict = {}

    # First, gather all CHARGE transactions from Payout Details sheets
    for payout in payouts:
        payout_id = payout['node']['id'][-10:]
        payout_sheet = workbook[f"Payout {payout_id}"]

        # Find where transaction details start (after the header row)
        for row_idx, row in enumerate(payout_sheet.iter_rows(min_row=1, values_only=True), 1):
            if row[0] == "Transaction Date":  # Found the header row
                transaction_start_row = row_idx + 1
                break

        # Collect all CHARGE transactions
        for row in payout_sheet.iter_rows(min_row=transaction_start_row, values_only=True):
            if row[1] == "CHARGE":  # Check Type column
                order_num = row[2]  # Order column
                amount = row[3]     # Amount column
                charges_dict[order_num] = amount

    # Now check Refund Details sheet for Auth Release refunds
    refund_sheet = workbook["Refund Details"]
    auth_releases = {}

    for row in refund_sheet.iter_rows(min_row=2, values_only=True):
        order_num = row[0]      # Order Number column
        if row[4] == "Auth Release":  # UPC column
            auth_releases[order_num] = row[3]  # Refund Amount(T) column

    # Create rows for Payout Comparison sheet
    for order_num, charge_amount in charges_dict.items():
        auth_release_amount = auth_releases.get(order_num, 0.00)
        total_amount = charge_amount + auth_release_amount

        comparison_sheet.append([
            order_num,
            charge_amount,
            auth_release_amount,
            total_amount
        ])

    # Format numbers in Payout Comparison sheet
    for row in comparison_sheet.iter_rows(min_row=2):
        for cell in row[1:]:  # Skip Order column
            cell.number_format = '#,##0.00'

    # Ensure export path exists
    if not os.path.exists(export_path):
        os.makedirs(export_path)

    # Save the Excel file with a timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file_with_timestamp = os.path.join(
        export_path, f"Shopify_Payouts_{start_date}_to_{end_date}_{timestamp}.xlsx")
    workbook.save(output_file_with_timestamp)
    print(f"Payout data exported to {output_file_with_timestamp}")


if __name__ == "__main__":
    payouts = fetch_payouts()
    output_file = f"Shopify_Payouts_{start_date}_to_{end_date}.xlsx"
    create_excel(payouts, output_file)
    print(f"Payout data exported to {output_file}")
